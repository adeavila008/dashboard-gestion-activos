"""
Extrae y filtra los datos reales (Direccion de Gestion de Activos) de los
Excel fuente, y los deja listos como JSON compacto para embeber en el
dashboard como dataset por defecto (el usuario puede reemplazarlo cargando un
Excel nuevo desde el navegador).

El IBReport financiero se arma escaneando TODOS los .xlsx de una carpeta
(tipicamente "ISES/Costos" en el OneDrive del usuario, donde van quedando
tanto los exports ya cerrados de años anteriores como el del año en curso,
que se reemplaza/actualiza mes a mes). No hace falta distinguir "historico"
de "actual": cada fila se identifica por numero de asiento + secuencia +
cuenta + periodo + proyecto + importe, y se deduplica automaticamente, asi
que no importa si un archivo nuevo vuelve a traer meses que ya estaban en
otro archivo procesado antes.

Uso:
    python3 extract_data.py --ibreport-dir "<carpeta con los .xlsx de Costos>" <ruta_indicadores_ga.xlsx>
    python3 extract_data.py <ruta_ibreport.xlsx> <ruta_indicadores_ga.xlsx>          (un solo archivo)
    python3 extract_data.py --ibreport-dir "<carpeta con los .xlsx de Costos>" --ibreport-only
        (solo actualiza el IBReport financiero; no toca default_baseline.json ni
        projects_registry.json — para la automatizacion semanal, que solo tiene
        acceso a la carpeta Costos, no al Excel "Indicadores GA" de Salud de
        Proyectos)

Modos "por partes" (para entornos con limite de tiempo por comando, ej. la
tarea programada semanal, donde procesar 5 Excel grandes de un tirón no
alcanza a caber en una sola llamada):
    python3 extract_data.py --cache-file "<un_excel.xlsx>" "<carpeta_cache>"
        Procesa UN solo Excel y guarda sus filas GA ya extraidas como JSON en
        <carpeta_cache>/<nombre_archivo>.json (rapido, pensado para correr un
        archivo por llamada).
    python3 extract_data.py --merge-cache "<carpeta_cache>"
        Junta todos los .json de <carpeta_cache> (uno por Excel, generados con
        --cache-file), deduplica y escribe data/default_ibreport.json. No toca
        baseline/projects_registry (equivalente a --ibreport-only).

Si no se pasan argumentos, busca los archivos por nombre dentro de
build/source/ (carpeta local, no versionada en git por defecto). Tambien se
soporta build/source/historical/*.xlsx (modo antiguo) si no se usa
--ibreport-dir.
"""
import json
import math
import os
import sys
import openpyxl
from datetime import datetime

BUILD_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_DIR = os.path.join(BUILD_DIR, "source")
HISTORICAL_DIR = os.path.join(SOURCE_DIR, "historical")
OUT = os.path.join(BUILD_DIR, "data")

DEFAULT_IBREPORT_NAME = "Consultoría - Inf. de ingresos, gastos y costos_Default (15).xlsx"
DEFAULT_BASELINE_NAME = "Indicadores GA _.xlsx"

_argv = sys.argv[1:]

CACHE_FILE_XLSX = None
CACHE_DIR = None
MERGE_CACHE_DIR = None
if len(_argv) >= 3 and _argv[0] == "--cache-file":
    CACHE_FILE_XLSX = _argv[1]
    CACHE_DIR = _argv[2]
elif len(_argv) >= 2 and _argv[0] == "--merge-cache":
    MERGE_CACHE_DIR = _argv[1]

IBREPORT_DIR = None
if len(_argv) >= 2 and _argv[0] == "--ibreport-dir":
    IBREPORT_DIR = _argv[1]
    _argv = _argv[2:]

IBREPORT_ONLY = "--ibreport-only" in _argv
if IBREPORT_ONLY:
    _argv = [a for a in _argv if a != "--ibreport-only"]

IBREPORT_PATH = _argv[0] if len(_argv) > 0 else os.path.join(SOURCE_DIR, DEFAULT_IBREPORT_NAME)
BASELINE_PATH = _argv[1] if len(_argv) > 1 else os.path.join(SOURCE_DIR, DEFAULT_BASELINE_NAME)

DIRECCION = "DIRECCION DE GESTIÓN DE ACTIVOS"

GA_PROJECT_CODES = {
    "1012",
    "AFN24100-101",
    "AFN24103-100",
    "AIR25103-100",
    "AIR25105-101",
    "CER24100-100",
    "ESA25101-100",
}


EXCEL_ERROR_TOKENS = {
    "#DIV/0!", "#N/A", "#VALUE!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#ERROR!",
}


def clean(v):
    if v is None:
        return None
    if isinstance(v, float):
        if math.isnan(v):
            return None
        return round(v, 4)
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str) and v.strip().upper() in EXCEL_ERROR_TOKENS:
        # celda con error de formula (ej. CPI = costo/0 en meses sin ejecucion aun)
        return None
    return v


IB_KEY_MAP = {
    "Eri_est (T)": "eriEst",
    "Direccion_linea (T)": "direccion",
    "Cuenta_mayor": "cuentaMayorCod",
    "Cuenta_mayor (T)": "cuentaMayor",
    "Cuenta": "cuentaCod",
    "Cuenta (T)": "cuenta",
    "Proyecto": "proyectoCod",
    "Proyecto (T)": "proyecto",
    "Año": "anio",
    "Periodo": "periodo",
    "Fcha. Asto.": "fecha",
    "Tipo de asientos": "tipo",
    "Número de asiento": "numAsiento",
    "Nº de secuencia": "secuencia",
    "Descripción": "descripcion",
    "Tercero (T)": "tercero",
    "Importe": "importe",
    "Empleado (T)": "empleado",
    "Activos Fijos": "activoFijoCod",
    "Activos Fijos (T)": "activoFijo",
}
IB_COLS = list(IB_KEY_MAP.values()) + ["esIngreso"]


def _find_ibreport_sheet(wb):
    for name in wb.sheetnames:
        if "ibreport" in name.lower():
            return name
    return wb.sheetnames[0]


def _extract_ibreport_rows(path):
    """Lee UN Excel de IBReport y devuelve solo las filas de la Direccion de
    Gestion de Activos, ya normalizadas segun IB_KEY_MAP/IB_COLS."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = _find_ibreport_sheet(wb)
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter)
    idx = {h: i for i, h in enumerate(headers)}
    i_dir = idx.get("Direccion_linea (T)")
    i_cmc = idx.get("Cuenta_mayor")
    if i_dir is None or i_cmc is None:
        raise ValueError(f"'{path}': no se encontraron las columnas esperadas de direccion/cuenta mayor.")

    out = []
    for row in rows_iter:
        if row[i_dir] != DIRECCION:
            continue
        rec = [clean(row[idx[src]]) if src in idx else None for src in IB_KEY_MAP]
        cod = str(row[i_cmc]) if row[i_cmc] is not None else ""
        rec.append(cod.startswith("4"))
        out.append(rec)
    wb.close()
    return out


def _dedup_key_indices():
    idx_key = {name: i for i, name in enumerate(IB_KEY_MAP.values())}
    return idx_key["numAsiento"], idx_key["secuencia"], idx_key["cuentaMayorCod"], idx_key["periodo"], idx_key["proyectoCod"], idx_key["importe"]


def _dedup_rows(rows_by_source):
    """rows_by_source: lista de (nombre_legible, lista_de_filas). Deduplica
    entre todas las fuentes usando la misma clave que extract_ibreport()."""
    i_numAsiento, i_secuencia, i_cuentaMayorCod, i_periodo, i_proyectoCod, i_importe = _dedup_key_indices()
    seen = set()
    all_rows = []
    for name, rows in rows_by_source:
        added = 0
        for rec in rows:
            key = (rec[i_numAsiento], rec[i_secuencia], rec[i_cuentaMayorCod], rec[i_periodo], rec[i_proyectoCod], rec[i_importe])
            if key in seen:
                continue
            seen.add(key)
            all_rows.append(rec)
            added += 1
        print(f"  {name}: {len(rows)} filas GA ({added} nuevas tras deduplicar)")
    return all_rows


def cache_one_file():
    """--cache-file <xlsx> <cache_dir>: procesa UN Excel y guarda sus filas GA
    ya extraidas (formato IB_COLS) como JSON en <cache_dir>. Pensado para
    correr un archivo por llamada cuando hay limite de tiempo por comando."""
    if not os.path.exists(CACHE_FILE_XLSX):
        raise FileNotFoundError(f"No existe el archivo '{CACHE_FILE_XLSX}'.")
    os.makedirs(CACHE_DIR, exist_ok=True)
    rows = _extract_ibreport_rows(CACHE_FILE_XLSX)
    base = os.path.splitext(os.path.basename(CACHE_FILE_XLSX))[0]
    out_path = os.path.join(CACHE_DIR, base + ".json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, separators=(",", ":"))
    print(f"Cache guardado: {out_path} ({len(rows)} filas GA)")


def merge_cache():
    """--merge-cache <cache_dir>: junta todos los .json generados con
    --cache-file, deduplica y escribe data/default_ibreport.json. Nunca toca
    baseline/projects_registry."""
    if not os.path.isdir(MERGE_CACHE_DIR):
        raise FileNotFoundError(f"No existe la carpeta de cache '{MERGE_CACHE_DIR}'.")
    files = sorted(fn for fn in os.listdir(MERGE_CACHE_DIR) if fn.lower().endswith(".json"))
    if not files:
        raise FileNotFoundError(f"No hay archivos .json en '{MERGE_CACHE_DIR}' (corre --cache-file primero para cada Excel).")
    rows_by_source = []
    for fn in files:
        with open(os.path.join(MERGE_CACHE_DIR, fn), encoding="utf-8") as f:
            rows_by_source.append((fn, json.load(f)))
    all_rows = _dedup_rows(rows_by_source)
    payload = {"cols": IB_COLS, "rows": all_rows}
    with open(f"{OUT}/default_ibreport.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    print("IBReport GA rows (total combinado desde cache):", len(all_rows))
    return all_rows, IB_COLS


def extract_ibreport():
    paths = []
    if IBREPORT_DIR:
        # Modo carpeta: procesa TODOS los .xlsx de la carpeta indicada (ej. la
        # carpeta "Costos" del usuario), sin distinguir historico/actual.
        if not os.path.isdir(IBREPORT_DIR):
            raise FileNotFoundError(f"No existe la carpeta --ibreport-dir '{IBREPORT_DIR}'.")
        for fn in sorted(os.listdir(IBREPORT_DIR)):
            if fn.lower().endswith((".xlsx", ".xls")) and not fn.startswith("~$"):
                paths.append(os.path.join(IBREPORT_DIR, fn))
    else:
        # Modo antiguo: historicos fijos (build/source/historical/*.xlsx), en
        # orden alfabetico (los nombres traen el rango de periodo, ej.
        # "202401 - 202406", asi que ordenan cronologicamente solos) + el
        # Excel "actual" pasado como primer argumento.
        if os.path.isdir(HISTORICAL_DIR):
            for fn in sorted(os.listdir(HISTORICAL_DIR)):
                if fn.lower().endswith((".xlsx", ".xls")):
                    paths.append(os.path.join(HISTORICAL_DIR, fn))
        if os.path.exists(IBREPORT_PATH):
            paths.append(IBREPORT_PATH)
        else:
            print(f"AVISO: no se encontro el IBReport actual en '{IBREPORT_PATH}'.")

    if not paths:
        raise FileNotFoundError("No se encontro ningun IBReport para procesar.")

    idx_key = {name: i for i, name in enumerate(IB_KEY_MAP.values())}
    i_numAsiento, i_secuencia = idx_key["numAsiento"], idx_key["secuencia"]
    i_cuentaMayorCod, i_periodo = idx_key["cuentaMayorCod"], idx_key["periodo"]
    i_proyectoCod, i_importe = idx_key["proyectoCod"], idx_key["importe"]

    seen = set()
    all_rows = []
    for path in paths:
        rows = _extract_ibreport_rows(path)
        added = 0
        for rec in rows:
            # dedup por si dos archivos llegaran a solaparse un mismo mes:
            # un asiento contable puntual queda identificado por su numero de
            # asiento + secuencia + cuenta + periodo + proyecto + importe.
            key = (rec[i_numAsiento], rec[i_secuencia], rec[i_cuentaMayorCod], rec[i_periodo], rec[i_proyectoCod], rec[i_importe])
            if key in seen:
                continue
            seen.add(key)
            all_rows.append(rec)
            added += 1
        print(f"  {os.path.basename(path)}: {len(rows)} filas GA ({added} nuevas tras deduplicar)")

    payload = {"cols": IB_COLS, "rows": all_rows}
    with open(f"{OUT}/default_ibreport.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    print("IBReport GA rows (total combinado, historico + actual):", len(all_rows))
    return all_rows, IB_COLS


def extract_baseline():
    path = BASELINE_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Consolidado"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers)}

    key_map = {
        "Gerencia": "gerencia", "Año": "anio", "Mes": "mes", "Estado": "estado",
        "Ceco": "cecoCod", "Proyecto": "proyectoNombre", "Cliente": "cliente",
        "Gerente": "gerente", "N° Contrato": "contrato", "Objeto": "objeto",
        "Fecha de Inicio": "fechaInicio", "Fecha Fin": "fechaFin",
        "Margen LB ": "margenLB", "Margena Ofertado": "margenOfertado",
        "Estado del Contrato": "estadoContrato",
        "Valor Contrato Inicial": "valorContratoInicial",
        "Valor Contrato Final": "valorContratoFinal",
        "Avance Planeado LB1 %": "avancePlanLB1", "Avance Planeado LB2 %": "avancePlanLB2",
        "Avance Planeado LB3 %": "avancePlanLB3", "Avance Real %": "avanceReal",
        "Variación Cronograma %": "variacionCronograma",
        "Cumplimiento Cronograma %": "cumplimientoCronograma",
        "WIP": "wip", "Saldo WIP": "saldoWip",
        "Presupuesto MCOP LB1": "presupuestoLB1", "Presupuesto MCOP LB2": "presupuestoLB2",
        "Presupuesto MCOP LB3": "presupuestoLB3",
        "Costo Planeado Mensual MCOP LB1": "costoPlanMensLB1",
        "Costo Planeado Acum MCOP LB1": "costoPlanAcumLB1",
        "Costo Planeado Mensual MCOP LB 2": "costoPlanMensLB2",
        "Costo Planeado Acum MCOP LB2": "costoPlanAcumLB2",
        "Costo Planeado Mensual MCOP LB 3": "costoPlanMensLB3",
        "Costo Planeado Acum MCOP LB3": "costoPlanAcumLB3",
        "Costo Real MCOP": "costoRealMens", "Costo Real Acum MCOP": "costoRealAcum",
        "Proyección Costos MCOP": "proyeccionCostoMens",
        "Proyección Costos Acum MCOP": "proyeccionCostoAcum",
        "Facturación MCOP LB1": "facturacionLB1", "Facturación MCOP LB2": "facturacionLB2",
        "Facturación MCOP LB3": "facturacionLB3",
        "Facturación Planeada Mensual MCOP LB1": "factPlanMensLB1",
        "Facturación Planeado Acum MCOP LB1": "factPlanAcumLB1",
        "Facturación Planeada Mensual MCOP LB2": "factPlanMensLB2",
        "Facturación Planeado Acum MCOP LB2": "factPlanAcumLB2",
        "Facturación Planeada Mensual MCOP LB3": "factPlanMensLB3",
        "Facturación Planeado Acum MCOP LB3": "factPlanAcumLB3",
        "Facturación Real MCOP": "factRealMens", "Facturación Real Acum MCOP": "factRealAcum",
        "Proyección Fact MCOP": "proyeccionFactMens", "Proyección Fact Acum MCOP": "proyeccionFactAcum",
        "Margen + WIP": "margenMasWip", "Margen Proyectado": "margenProyectado",
        "CV (Variación Costo) MCOP": "cv", "CV (% Variación Costo) MCOP": "cvPct",
        "EV (Valor ganado) MCOP": "ev", "CPI (Cost performance index)": "cpi",
        "SPI (Schedule performance index)": "spi", "EAC (Estimated at completion)": "eac",
        "ETC (Estimate to complete)": "etc", "CSI (Cost schedule index)": "csi",
        "Hechos Relevantes": "hechosRelevantes", "Proxímos Pasos": "proximosPasos",
        "Comentarios Avance": "comentariosAvance",
        "Comentarios Facturación": "comentariosFacturacion",
        "Comentarios Costos": "comentariosCostos", "Margen real": "margenReal",
    }

    out = []
    for row in rows[1:]:
        ceco = row[idx["Ceco"]]
        if ceco not in GA_PROJECT_CODES:
            continue
        rec = {}
        for src, dst in key_map.items():
            rec[dst] = clean(row[idx[src]])
        rec["cecoCod"] = ceco
        out.append(rec)

    with open(f"{OUT}/default_baseline.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    print("Baseline GA rows:", len(out))
    return out


def build_projects_registry(baseline_rows):
    registry = {}
    for r in baseline_rows:
        code = r["cecoCod"]
        if code == "1012":
            continue
        if code not in registry:
            registry[code] = {
                "codigo": code, "nombre": r.get("proyectoNombre"),
                "cliente": r.get("cliente"), "gerente": r.get("gerente"),
                "contrato": r.get("contrato"), "objeto": r.get("objeto"),
                "estado": r.get("estado"), "fechaInicio": r.get("fechaInicio"),
                "fechaFin": r.get("fechaFin"),
                "valorContratoInicial": r.get("valorContratoInicial"),
                "margenLB": r.get("margenLB"),
            }
    with open(f"{OUT}/projects_registry.json", "w", encoding="utf-8") as f:
        json.dump(list(registry.values()), f, ensure_ascii=False, separators=(",", ":"))
    print("Proyectos registrados:", len(registry))


if __name__ == "__main__":
    if CACHE_FILE_XLSX is not None:
        cache_one_file()
    elif MERGE_CACHE_DIR is not None:
        merge_cache()
    else:
        extract_ibreport()
        if not IBREPORT_ONLY:
            baseline = extract_baseline()
            build_projects_registry(baseline)
        else:
            print("Modo --ibreport-only: no se tocan default_baseline.json ni projects_registry.json.")
